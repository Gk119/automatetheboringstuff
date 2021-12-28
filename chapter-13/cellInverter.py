import openpyxl

filename = input("Enter the path and name of the file: ")

wb = openpyxl.load_workbook(filename)
sheet = wb.active

wb2 = openpyxl.Workbook()
sheet2 = wb2.active

for i in range(1,sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        sheet2.cell(row= j, column= i).value = sheet.cell(row= i, column= j).value

wb2.save(filename.replace('.xlsx','') + 'inverted.xlsx')