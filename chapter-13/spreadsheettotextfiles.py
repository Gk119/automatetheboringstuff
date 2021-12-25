import openpyxl, sys

wb = openpyxl.load_workbook(sys.argv[1])
sheet = wb.active

for col in range(1,sheet.max_column + 1):
    filename = 'File' + str(col) + '.txt'
    file = open(filename,'w')
    content = ""
    for row in range(1,sheet.max_row + 1):
        content += sheet.cell(row= row, column= col).value

    print('File saved: ' + filename)
    file.write(content)
    file.close()

      