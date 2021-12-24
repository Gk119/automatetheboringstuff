import openpyxl, sys
from openpyxl.styles import Font

workbook = openpyxl.Workbook()
sheet = workbook.active

N = int(sys.argv[1])
print(N)
boldFont = Font(bold= True)
for i in range(1,N+1):
    sheet.cell(row= 1, column= i + 1).value = str(i)
    sheet.cell(row= 1, column= i + 1).font = boldFont

for i in range(1,N+1):
    sheet.cell(row= i + 1, column= 1).value = str(i)
    sheet.cell(row= i + 1, column= 1).font = boldFont

for i in range(1,N+1):
    for j in range(1,N+1):
        sheet.cell(row= i + 1, column= j + 1).value = i*j

workbook.save('MultiplicationTable%sx%s.xlsx'%(N,N))

