import openpyxl, sys

wb = openpyxl.load_workbook(sys.argv[3])
N = int(sys.argv[1])
M = int(sys.argv[2])

sheet = wb.active

max_row = sheet.max_row
max_column = sheet.max_column

wb2 = openpyxl.Workbook()
sheet2 = wb2.active

for i in range(1,N):
    for j in range(1,max_column + 1):
        sheet2.cell(row= i, column= j).value = sheet.cell(row= i, column= j).value

for i in range(N + M, max_row + M + 1):
    for j in range(1,max_column + 1):
        sheet2.cell(row= i, column= j).value = sheet.cell(row= i - M, column= j).value

wb2.save('tst.xlsx')


